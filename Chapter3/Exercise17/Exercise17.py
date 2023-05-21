from openpyxl import Workbook
from math import sin
wb = Workbook()
ws = wb.active
ws.title = "MeineErsteTabelle"
ws['A1'] = "Wertebereich"
ws['B1'] = "f(x)=x*sin(x)"
count = 2
for x in range(-31,32,1):
    x = x/10.0
    ws['A'+str(count)] = x
    ws['B'+str(count)] = x*sin(x)
    count += 1
wb.save("Datei17.xlsx")